import pandas as pd
from io import BytesIO
from django.utils import timezone
from propertydata.models import *
from django.db.models import Q
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import json
class CustomExportResource:
    """
    Dynamically generates Excel data for a given CustomExportOptions instance.
    Includes up to 4 related contacts with wireless & landline numbers.
    """
    def __init__(self, export_option, queryset, list_name):
        self.export_option = export_option
        self.queryset = queryset
        self.listname = list_name

    def dehydrate_foreclosure(self, obj):
        """
        Convert a Foreclosure instance into a dict with up to 4 contact groups.
        """
        base = {
            "State": obj.state or "-",
            "County": obj.county or "-",
            "Case Number": obj.case_number or "-",
            "Sale Date": obj.sale_date or "-",
            "Sale Type": obj.sale_type or "-",
            "Judgment Amount": obj.fcl_final_judgment,
            "Sale Price": obj.sale_price,
            "Possible Surplus": obj.possible_surplus,
            "Verified Surplus": obj.verified_surplus,
        }

        # --- Handle Plaintiff (ManyToMany) -
        plaintiffs = obj.plaintiff.all() if hasattr(obj, "plaintiff") else []
        if plaintiffs:
            base["Plaintiff"] = ", ".join([str(p) for p in plaintiffs if p]) or "-"
        else:
            base["Plaintiff"] = "-"

        # --- Handle Defendant (ManyToMany) -
        defendants = obj.defendant.all() if hasattr(obj, "defendant") else []
        if defendants:
            base["Defendant"] = ", ".join([str(d) for d in defendants if d]) or "-"
        else:
            base["Defendant"] = "-"

        # --- Handle related property (ManyToMany) -
        STATE_ABBREVIATIONS = {
            "Alabama": "AL",
            "Alaska": "AK",
            "Arizona": "AZ",
            "Arkansas": "AR",
            "California": "CA",
            "Colorado": "CO",
            "Connecticut": "CT",
            "Delaware": "DE",
            "Florida": "FL",
            "Georgia": "GA",
            "Hawaii": "HI",
            "Idaho": "ID",
            "Illinois": "IL",
            "Indiana": "IN",
            "Iowa": "IA",
            "Kansas": "KS",
            "Kentucky": "KY",
            "Louisiana": "LA",
            "Maine": "ME",
            "Maryland": "MD",
            "Massachusetts": "MA",
            "Michigan": "MI",
            "Minnesota": "MN",
            "Mississippi": "MS",
            "Missouri": "MO",
            "Montana": "MT",
            "Nebraska": "NE",
            "Nevada": "NV",
            "New Hampshire": "NH",
            "New Jersey": "NJ",
            "New Mexico": "NM",
            "New York": "NY",
            "North Carolina": "NC",
            "North Dakota": "ND",
            "Ohio": "OH",
            "Oklahoma": "OK",
            "Oregon": "OR",
            "Pennsylvania": "PA",
            "Rhode Island": "RI",
            "South Carolina": "SC",
            "South Dakota": "SD",
            "Tennessee": "TN",
            "Texas": "TX",
            "Utah": "UT",
            "Vermont": "VT",
            "Virginia": "VA",
            "Washington": "WA",
            "West Virginia": "WV",
            "Wisconsin": "WI",
            "Wyoming": "WY",
            "District of Columbia": "DC",
        }
        first_property = obj.property.first() if hasattr(obj, "property") else None
        if first_property:
            state_name = first_property.state or "-"
            state_short = STATE_ABBREVIATIONS.get(state_name.strip().title(), "-")
            base["Street Address"] = first_property.street_address or "-"
            base["City"] = first_property.city or "-"
            base["ST"] = state_short
            base["Zip"] = first_property.zip_code or "-"
            base["Parcel"] = first_property.parcel or "-"
        else:
            base["Street Address"] = "-"
            base["City"] = "-"
            base["ST"] = "-"
            base["Zip"] = "-"
            base["Parcel"] = "-"

        # --- Add up to 5 contacts (defendants + related contacts) ---
        contacts = list(obj.defendant.all())

        # include related contacts of each defendant
        for d in obj.defendant.all():
            for rc in d.related_contacts.all():
                if rc not in contacts:
                    contacts.append(rc)

        # limit total to 5
        contacts = contacts[:5]

        for i in range(5):
            cnum = i + 1
            if i < len(contacts):
                contact = contacts[i]
                
                base[f"Contact Name {cnum}"] = str(contact) or "-"
                # pick up to 1 primary email or first 4 if you want more
                email_list = list(contact.emails.values_list("email_address", flat=True)[:4])
                base[f"Email {cnum}"] = ", ".join(email_list) if email_list else "-"

                # wireless numbers (up to 4)
                wireless_list = list(contact.wireless.values_list("w_number", flat=True)[:4])
                for j in range(4):
                    base[f"Wireless {cnum}.{j+1}"] = wireless_list[j] if j < len(wireless_list) else "-"

                # landline numbers (up to 4)
                landline_list = list(contact.landline.values_list("l_number", flat=True)[:4])
                for j in range(4):
                    base[f"Landline {cnum}.{j+1}"] = landline_list[j] if j < len(landline_list) else "-"
            else:
                # fill empty contact slots with "-"
                base[f"Contact Name {cnum}"] = "-"
                base[f"Email {cnum}"] = "-"
                for j in range(4):
                    base[f"Wireless {cnum}.{j+1}"] = "-"
                    base[f"Landline {cnum}.{j+1}"] = "-"

        return base

    def to_dataframe(self):
        queryset = self.queryset
        data = [self.dehydrate_foreclosure(obj) for obj in queryset]
        df = pd.DataFrame(data)

        hardcoded_columns = [
            "State",
            "County",
            "Case Number",
            "Sale Date",
            "Sale Type",
            "Judgment Amount",
            "Sale Price",
            "Possible Surplus",
            "Verified Surplus",

            "Plaintiff",

            "Defendant",

            "Parcel",
            "Street Address",
            "City",
            "ST",
            "Zip",

            "Contact Name 1",
            "Email 1",
            "Wireless 1.1",
            "Wireless 1.2",
            "Wireless 1.3",
            "Wireless 1.4",
            "Landline 1.1",
            "Landline 1.2",
            "Landline 1.3",
            "Landline 1.4",
            "Contact Name 2",

            "Email 2",
            "Wireless 2.1",
            "Wireless 2.2",
            "Wireless 2.3",
            "Wireless 2.4",
            "Landline 2.1",
            "Landline 2.2",
            "Landline 2.3",
            "Landline 2.4",

            "Contact Name 3",
            "Email 3",
            "Wireless 3.1",
            "Wireless 3.2",
            "Wireless 3.3",
            "Wireless 3.4",
            "Landline 3.1",
            "Landline 3.2",
            "Landline 3.3",
            "Landline 3.4",

            "Contact Name 4",
            "Email 4",
            "Wireless 4.1",
            "Wireless 4.2",
            "Wireless 4.3",
            "Wireless 4.4",
            "Landline 4.1",
            "Landline 4.2",
            "Landline 4.3",
            "Landline 4.4",

            "Contact Name 5",
            "Email 5",
            "Wireless 5.1",
            "Wireless 5.2",
            "Wireless 5.3",
            "Wireless 5.4",
            "Landline 5.1",
            "Landline 5.2",
            "Landline 5.3",
            "Landline 5.4",
        ]
        
        if df.empty:
            print("[⚠️ WARNING] Dataframe is empty — skipping column filtering")
            return df
        available_cols = list(df.columns)
        
        matched_cols = [c for c in hardcoded_columns if c in available_cols]
        if matched_cols:
            df = df[matched_cols]
        else:
            print(f"[⚠️ WARNING] No matching columns found, exporting all columns")
            print("Available:", available_cols)
        # ✅ Fix timezone-aware datetimes before Excel export
        for col in df.select_dtypes(include=["datetimetz"]).columns:
            df[col] = df[col].dt.tz_localize(None)
        
        return df

    def export_to_excel(self):
        df = self.to_dataframe()
        buffer = BytesIO()
        filename = f"{timezone.now().date()} {self.listname} List.xlsx"
        # ✅ Ensure currency columns are numeric
        number_data = [
            "Judgment Amount",
            "Sale Price",
            "Possible Surplus",
            "Verified Surplus",
            "Zip"
        ]

        for col in number_data:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        
        text_upper_columns = [
            "State",
            "County",
            "Sale Type"
        ]
        for col in text_upper_columns:
            if col in df.columns:
                df[col] = df[col].fillna("").str.upper()

        text_proper_columns = [
            "Plaintiff",
            "Defendant",
            "Contact Name 1",
            "Contact Name 2",
            "Contact Name 3",
            "Contact Name 4",
            "Contact Name 5",
        ]
        for col in text_proper_columns:
            if col in df.columns:
                df[col] = df[col].fillna("").str.title()

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            sheet_name = f"{self.listname} Data"
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            # --- Style header after writing ---
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Header styling
            header_fill = PatternFill(start_color="516699", end_color="516699", fill_type="solid")  # blue background
            header_font = Font(color="FFFFFF", bold=True, size=12)  # white bold text
            header_align = Alignment(horizontal="center", vertical="center")

            # Apply style to all header cells
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                # Optional: Auto-fit column width based on content
                max_length = max(
                    (len(str(cell_value)) for cell_value in [column_title] + df[column_title].astype(str).tolist()),
                    default=0
                )
                worksheet.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 100)  # cap width
            # date and currency formats
            currency_format = '"$"#,##0.00'
            date_format = "MM/DD/YYYY"

            currency_columns = [
                "Judgment Amount",
                "Sale Price",
                "Possible Surplus",
                "Verified Surplus"
            ]

            date_columns = [
                "Sale Date"
            ]

            for col_num, column_title in enumerate(df.columns, 1):

                # Currency columns
                if column_title in currency_columns:
                    for row in range(2, len(df) + 2):
                        worksheet.cell(row=row, column=col_num).number_format = currency_format

                # Date columns
                if column_title in date_columns:
                    for row in range(2, len(df) + 2):
                        worksheet.cell(row=row, column=col_num).number_format = date_format

        buffer.seek(0)
        return filename, buffer, df