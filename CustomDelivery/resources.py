import pandas as pd
from io import BytesIO
from django.utils import timezone
from propertydata.models import *
from django.db.models import Q
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import json
class CustomExportResource:
    """
    Dynamically generates Excel data for a given CustomExportOptions instance.
    Includes up to 4 related contacts with wireless & landline numbers.
    """
    def __init__(self, export_option):
        self.export_option = export_option

    def get_queryset(self):
        """
        Build a queryset that matches each ExportLeadFilter exactly.
        Avoids mixing states/types between filters.
        """
        filters = self.export_option.filter_option.all()
        q_objects = Q()

        for f in filters:
            q = Q(state=f.state)
            if f.sale_type:
                q &= Q(sale_type=f.sale_type)
            if f.sale_status:
                q &= Q(sale_status=f.sale_status)
            if f.surplus_status:
                q &= Q(surplus_status=f.surplus_status)
            q_objects |= q  # OR between filters, but AND inside each

        queryset = Foreclosure.objects.filter(q_objects).distinct()
        queryset = queryset.filter(published=True)
        # Omit already delivered or old leads
        queryset = self.exclude_delivered_and_old(queryset)

        return queryset

    def exclude_delivered_and_old(self, queryset):
        """
        Remove old_leads and already delivered leads based on delivery_type.
        """
        export = self.export_option

        # Exclude old leads (always)
        #queryset = queryset.exclude(pk__in=export.old_leads.values_list("pk", flat=True))
        queryset = queryset.exclude(pk__in=export.client.old_leads.values_list("pk", flat=True))

        # Exclude previously delivered ones of the same type
        if export.delivery_type == "pre-foreclosure":
            #queryset = queryset.exclude(pk__in=export.pre_foreclosure.values_list("pk", flat=True))
            queryset = queryset.exclude(pk__in=export.client.pre_foreclosure.values_list("pk", flat=True))
        elif export.delivery_type == "post-foreclosure":
            #queryset = queryset.exclude(pk__in=export.post_foreclosure.values_list("pk", flat=True))
            queryset = queryset.exclude(pk__in=export.client.post_foreclosure.values_list("pk", flat=True))
        elif export.delivery_type == "verified":
            #queryset = queryset.exclude(pk__in=export.verified_surplus.values_list("pk", flat=True))
            queryset = queryset.exclude(pk__in=export.client.verified_surplus.values_list("pk", flat=True))

        return queryset

    def dehydrate_foreclosure(self, obj):
        """
        Convert a Foreclosure instance into a dict with up to 4 contact groups.
        """
        base = {
            "ID": obj.id,
            "State": obj.state or "-",
            "County": obj.county or "-",
            "Sale Type": obj.sale_type or "-",
            "Sale Status": obj.sale_status or "-",
            "Surplus Status": obj.surplus_status or "-",
            "Case Number": obj.case_number or "-",
            "Sale Date": obj.sale_date or "-",
            "Judgment Amount": obj.fcl_final_judgment or "-",
            "Sale Price": obj.sale_price or "-",
            "Possible Surplus": obj.possible_surplus or "-",
            "Verified Surplus": obj.verified_surplus or "-",
            "Changed At":obj.changed_at or "-",
            "Lead ID":"",
            "Confirmation Date":"",
            "Foreclosure Deed":"-",
            "Prior Deed":"-",
            "Mailing Address":"-",          
        }

        # --- Handle Plaintiff (ManyToMany) -
        plaintiffs = obj.plaintiff.all() if hasattr(obj, "plaintiff") else []
        if plaintiffs:
            base["Plaintiff"] = ", ".join([str(p) for p in plaintiffs if p]) or "-"
        else:
            base["Plaintiff"] = "-"

        # --- Handle Defendant (ManyToMany) -
        defendants = obj.defendant.all() if hasattr(obj, "defendant") else []
        if defendants:
            first_defendant = defendants[0]
            other_defendants = defendants[1:]
            # Default placeholders
            base["First Name"] = "-"
            base["Middle Name"] = "-"
            base["Last Name"] = "-"
            base["Additional Defendants"] = "-"
            # Check for business name first
            if first_defendant.business_name:
                base["First Name"] = str(first_defendant).strip()
            else:
                base["First Name"] = first_defendant.first_name.strip() if first_defendant.first_name else "-"
                base["Middle Name"] = first_defendant.middle_name.strip() if first_defendant.middle_name else "-"
                base["Last Name"] = first_defendant.last_name.strip() if first_defendant.last_name else "-"
            # Combine remaining defendants into one string
            if other_defendants:
                base["Additional Defendants"] = ", ".join([str(d) for d in other_defendants if d]) or "-"
            base["Defendant"] = ", ".join([str(d) for d in defendants if d]) or "-"
        else:
            base["First Name"] = "-"
            base["Middle Name"] = "-"
            base["Last Name"] = "-"
            base["Additional Defendants"] = "-"
            base["Defendant"] = "-"

        # --- Handle related property (ManyToMany) -
        STATE_ABBREVIATIONS = {
            "Alabama": "AL",
            "Alaska": "AK",
            "Arizona": "AZ",
            "Arkansas": "AR",
            "California": "CA",
            "Colorado": "CO",
            "Connecticut": "CT",
            "Delaware": "DE",
            "Florida": "FL",
            "Georgia": "GA",
            "Hawaii": "HI",
            "Idaho": "ID",
            "Illinois": "IL",
            "Indiana": "IN",
            "Iowa": "IA",
            "Kansas": "KS",
            "Kentucky": "KY",
            "Louisiana": "LA",
            "Maine": "ME",
            "Maryland": "MD",
            "Massachusetts": "MA",
            "Michigan": "MI",
            "Minnesota": "MN",
            "Mississippi": "MS",
            "Missouri": "MO",
            "Montana": "MT",
            "Nebraska": "NE",
            "Nevada": "NV",
            "New Hampshire": "NH",
            "New Jersey": "NJ",
            "New Mexico": "NM",
            "New York": "NY",
            "North Carolina": "NC",
            "North Dakota": "ND",
            "Ohio": "OH",
            "Oklahoma": "OK",
            "Oregon": "OR",
            "Pennsylvania": "PA",
            "Rhode Island": "RI",
            "South Carolina": "SC",
            "South Dakota": "SD",
            "Tennessee": "TN",
            "Texas": "TX",
            "Utah": "UT",
            "Vermont": "VT",
            "Virginia": "VA",
            "Washington": "WA",
            "West Virginia": "WV",
            "Wisconsin": "WI",
            "Wyoming": "WY",
            "District of Columbia": "DC",
        }
        first_property = obj.property.first() if hasattr(obj, "property") else None
        if first_property:
            state_name = first_property.state or "-"
            state_short = STATE_ABBREVIATIONS.get(state_name.strip().title(), "-")
            base["Street Address"] = first_property.street_address or "-"
            base["City"] = first_property.city or "-"
            base["ST"] = state_short
            base["Zip"] = first_property.zip_code or "-"
            base["Parcel"] = first_property.parcel or "-"
        else:
            base["Street Address"] = "-"
            base["City"] = "-"
            base["ST"] = "-"
            base["Zip"] = "-"
            base["Parcel"] = "-"

        # --- Add up to 5 contacts (defendants + related contacts) ---
        contacts = list(obj.defendant.all())

        # include related contacts of each defendant
        for d in obj.defendant.all():
            for rc in d.related_contacts.all():
                if rc not in contacts:
                    contacts.append(rc)

        # limit total to 5
        contacts = contacts[:5]

        for i in range(5):
            cnum = i + 1
            if i < len(contacts):
                contact = contacts[i]
                
                base[f"Contact Name {cnum}"] = str(contact) or "-"
                # pick up to 1 primary email or first 4 if you want more
                email_list = list(contact.emails.values_list("email_address", flat=True)[:4])
                base[f"Email {cnum}"] = ", ".join(email_list) if email_list else "-"

                # wireless numbers (up to 4)
                wireless_list = list(contact.wireless.values_list("w_number", flat=True)[:4])
                for j in range(4):
                    base[f"Wireless {cnum}.{j+1}"] = wireless_list[j] if j < len(wireless_list) else "-"

                # landline numbers (up to 4)
                landline_list = list(contact.landline.values_list("l_number", flat=True)[:4])
                for j in range(4):
                    base[f"Landline {cnum}.{j+1}"] = landline_list[j] if j < len(landline_list) else "-"
            else:
                # fill empty contact slots with "-"
                base[f"Contact Name {cnum}"] = "-"
                base[f"Email {cnum}"] = "-"
                for j in range(4):
                    base[f"Wireless {cnum}.{j+1}"] = "-"
                    base[f"Landline {cnum}.{j+1}"] = "-"

        return base

    def dehydrate_foreclosure_vertical(self, obj):
        """
        Generate multiple rows for each contact (defendants + related contacts),
        each row containing foreclosure info + single contact info.
        """
        base_info = {
            "ID": obj.id,
            "State": obj.state or "-",
            "County": obj.county or "-",
            "Sale Type": obj.sale_type or "-",
            "Sale Status": obj.sale_status or "-",
            "Surplus Status": obj.surplus_status or "-",
            "Case Number": obj.case_number or "-",
            "Sale Date": obj.sale_date or "-",
            "Possible Surplus": obj.possible_surplus or "-",
            "Verified Surplus": obj.verified_surplus or "-",
            "Judgment Amount": obj.fcl_final_judgment or "-",
            "Sale Price": obj.sale_price or "-",
            "Changed At":obj.changed_at or "-",
            "Lead ID":"",
            "Confirmation Date":"",
            "Foreclosure Deed":"-",
            "Prior Deed":"-",
            "Mailing Address":"-",
        }

        # --- Handle Plaintiff (ManyToMany) -
        plaintiffs = obj.plaintiff.all() if hasattr(obj, "plaintiff") else []
        if plaintiffs:
            base_info["Plaintiff"] = ", ".join([str(p) for p in plaintiffs if p]) or "-"
        else:
            base_info["Plaintiff"] = "-"

        # --- Handle Defendant (ManyToMany) -
        defendants = obj.defendant.all() if hasattr(obj, "defendant") else []
        if defendants:
            first_defendant = defendants[0]
            other_defendants = defendants[1:]
            # Default placeholders
            base_info["First Name"] = "-"
            base_info["Middle Name"] = "-"
            base_info["Last Name"] = "-"
            base_info["Additional Defendants"] = "-"
            # Check for business name first
            if first_defendant.business_name:
                base_info["First Name"] = str(first_defendant).strip()
            else:
                base_info["First Name"] = first_defendant.first_name.strip() if first_defendant.first_name else "-"
                base_info["Middle Name"] = first_defendant.middle_name.strip() if first_defendant.middle_name else "-"
                base_info["Last Name"] = first_defendant.last_name.strip() if first_defendant.last_name else "-"
            # Combine remaining defendants into one string
            if other_defendants:
                base_info["Additional Defendants"] = ", ".join([str(d) for d in other_defendants if d]) or "-"
            base_info["Defendant"] = ", ".join([str(d) for d in defendants if d]) or "-"
        else:
            base_info["First Name"] = "-"
            base_info["Middle Name"] = "-"
            base_info["Last Name"] = "-"
            base_info["Additional Defendants"] = "-"
            base_info["Defendant"] = "-"

        # --- Handle related property (ManyToMany) -
        STATE_ABBREVIATIONS = {
            "Alabama": "AL",
            "Alaska": "AK",
            "Arizona": "AZ",
            "Arkansas": "AR",
            "California": "CA",
            "Colorado": "CO",
            "Connecticut": "CT",
            "Delaware": "DE",
            "Florida": "FL",
            "Georgia": "GA",
            "Hawaii": "HI",
            "Idaho": "ID",
            "Illinois": "IL",
            "Indiana": "IN",
            "Iowa": "IA",
            "Kansas": "KS",
            "Kentucky": "KY",
            "Louisiana": "LA",
            "Maine": "ME",
            "Maryland": "MD",
            "Massachusetts": "MA",
            "Michigan": "MI",
            "Minnesota": "MN",
            "Mississippi": "MS",
            "Missouri": "MO",
            "Montana": "MT",
            "Nebraska": "NE",
            "Nevada": "NV",
            "New Hampshire": "NH",
            "New Jersey": "NJ",
            "New Mexico": "NM",
            "New York": "NY",
            "North Carolina": "NC",
            "North Dakota": "ND",
            "Ohio": "OH",
            "Oklahoma": "OK",
            "Oregon": "OR",
            "Pennsylvania": "PA",
            "Rhode Island": "RI",
            "South Carolina": "SC",
            "South Dakota": "SD",
            "Tennessee": "TN",
            "Texas": "TX",
            "Utah": "UT",
            "Vermont": "VT",
            "Virginia": "VA",
            "Washington": "WA",
            "West Virginia": "WV",
            "Wisconsin": "WI",
            "Wyoming": "WY",
            "District of Columbia": "DC",
        }

        # Handle related property
        first_property = obj.property.first() if hasattr(obj, "property") else None
        if first_property:
            state_name = first_property.state or "-"
            state_short = STATE_ABBREVIATIONS.get(state_name.strip().title(), "-")
            base_info.update({
                "Street Address": first_property.street_address or "-",
                "City": first_property.city or "-",
                "ST": state_short,
                "Zip": first_property.zip_code or "-",
                "Parcel": first_property.parcel or "-"
            })
        else:
            base_info.update({
                "Street Address": "-",
                "City": "-",
                "ST": "-",
                "Zip": "-",
                "Parcel": "-"
            })

        # Combine all contacts (defendants + related)
        contacts = list(obj.defendant.all())
        for d in obj.defendant.all():
            for rc in d.related_contacts.all():
                if rc not in contacts:
                    contacts.append(rc)

        contacts = contacts[:5]  # limit to 5

        rows = []
        if not contacts:
            # still create one row with blanks if no contact
            base_copy = base_info.copy()
            base_copy.update({
                "Contact Name": "-",
                "Email": "-",
                "Wireless 1": "-",
                "Wireless 2": "-",
                "Wireless 3": "-",
                "Wireless 4": "-",
                "Landline 1": "-",
                "Landline 2": "-",
                "Landline 3": "-",
                "Landline 4": "-",
            })
            rows.append(base_copy)
        else:
            for contact in contacts:
                contact_data = base_info.copy()
                contact_data["Contact Name"] = str(contact) or "-"

                # Emails
                email_list = list(contact.emails.values_list("email_address", flat=True)[:4])
                contact_data["Email"] = ", ".join(email_list) if email_list else "-"

                # Wireless
                wireless_list = list(contact.wireless.values_list("w_number", flat=True)[:4])
                #contact_data["Wireless"] = ", ".join(wireless_list) if wireless_list else "-"
                for j in range(4):
                    contact_data[f"Wireless {j+1}"] = wireless_list[j] if j < len(wireless_list) else "-"


                # Landline
                landline_list = list(contact.landline.values_list("l_number", flat=True)[:4])
                #contact_data["Landline"] = ", ".join(landline_list) if landline_list else "-"
                for j in range(4):
                    contact_data[f"Landline {j+1}"] = landline_list[j] if j < len(landline_list) else "-"

                rows.append(contact_data)

        return rows
    def to_dataframe(self):
        queryset = self.get_queryset()

        # Choose layout mode (horizontal or vertical)
        if getattr(self.export_option.client, "contact_align", "horizontal") == "vertical":
            data = []
            for obj in queryset:
                data.extend(self.dehydrate_foreclosure_vertical(obj))
            df = pd.DataFrame(data)
        else:
            data = [self.dehydrate_foreclosure(obj) for obj in queryset]
            df = pd.DataFrame(data)

        # 🩵 Safe filtering by client-specified columns
        client_columns = getattr(self.export_option.client, "columns", None)
        # 🧩 Convert JSON string to list if necessary
        if isinstance(client_columns, str):
            try:
                client_columns = json.loads(client_columns)
            except json.JSONDecodeError:
                print(f"[⚠️ WARNING] Failed to decode client columns for {self.export_option.client.name}: {client_columns}")
                client_columns = []
        if client_columns:
            available_cols = list(df.columns)
            matched_cols = [c for c in client_columns if c in available_cols]

            # Debug print to help identify mismatches
            print(f"\n--- Column Debug for {self.export_option.client.name} ---")
            print(f"Requested Columns: {client_columns}")
            print(f"Available Columns: {available_cols[:10]} ... ({len(available_cols)} total)")
            print(f"Matched Columns: {matched_cols}\n")
            for req in client_columns:
                if req not in available_cols:
                    print(f"❌ No match for: '{req}'")
                    close = [col for col in available_cols if col.lower().strip() == req.lower().strip()]
                    if close:
                        print(f"   🔸 Possible close match: {close}")
            if matched_cols:
                df = df[matched_cols]
            else:
                print(f"[⚠️ WARNING] No matching columns found for {self.export_option.client.name}. Exporting all columns instead.")
        else:
            print(f"[ℹ️ INFO] No custom columns specified for {self.export_option.client.name}. Exporting all columns.")

        # ✅ Fix timezone-aware datetimes before Excel export
        for col in df.select_dtypes(include=["datetimetz"]).columns:
            df[col] = df[col].dt.tz_localize(None)

        return df


    def export_to_excel(self):
        df = self.to_dataframe()
        buffer = BytesIO()
        filename = f"{timezone.now().date()} {self.export_option.delivery_type} List - {self.export_option.client.name}.xlsx"

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            sheet_name = f"{self.export_option.delivery_type.title()} Leads"
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            # --- Style header after writing ---
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Header styling
            header_fill = PatternFill(start_color="516699", end_color="516699", fill_type="solid")  # blue background
            header_font = Font(color="FFFFFF", bold=True, size=12)  # white bold text
            header_align = Alignment(horizontal="center", vertical="center")

            # Apply style to all header cells
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                # Optional: Auto-fit column width based on content
                max_length = max(
                    (len(str(cell_value)) for cell_value in [column_title] + df[column_title].astype(str).tolist()),
                    default=0
                )
                worksheet.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 100)  # cap width


        buffer.seek(0)
        return filename, buffer, df