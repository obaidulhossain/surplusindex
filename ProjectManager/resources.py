import pandas as pd
from io import BytesIO
from django.utils import timezone
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


class DashboardCloneExportResource:
    """
    Dashboard-only export resource.
    - Independent clone (no dependency on CustomExportResource)
    - Fixed field list & order
    - Queryset injected from view
    """

    # 🔐 EXACT column order (as provided)
    FIELDS = [
        "ID", "Changed At", "Surplus Status", "Sale Status", "State", "County",
        "Case Number", "Sale Date", "Confirmation Date", "Sale Type",
        "Plaintiff", "Defendant",
        "Judgment Amount", "Sale Price", "Possible Surplus", "Verified Surplus",
        "Parcel", "Street Address", "City", "ST", "Zip", "Mailing Address",

        "Contact Name 1", "Email 1",
        "Wireless 1.1", "Wireless 1.2", "Wireless 1.3", "Wireless 1.4",
        "Landline 1.1", "Landline 1.2", "Landline 1.3", "Landline 1.4",

        "Contact Name 2", "Email 2",
        "Wireless 2.1", "Wireless 2.2", "Wireless 2.3", "Wireless 2.4",
        "Landline 2.1", "Landline 2.2", "Landline 2.3", "Landline 2.4",

        "Contact Name 3", "Email 3",
        "Wireless 3.1", "Wireless 3.2", "Wireless 3.3", "Wireless 3.4",
        "Landline 3.1", "Landline 3.2", "Landline 3.3", "Landline 3.4",

        "Contact Name 4", "Email 4",
        "Wireless 4.1", "Wireless 4.2", "Wireless 4.3", "Wireless 4.4",
        "Landline 4.1", "Landline 4.2", "Landline 4.3", "Landline 4.4",

        "Contact Name 5", "Email 5",
        "Wireless 5.1", "Wireless 5.2", "Wireless 5.3", "Wireless 5.4",
        "Landline 5.1", "Landline 5.2", "Landline 5.3", "Landline 5.4",
    ]

    STATE_ABBREVIATIONS = {
        "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
        "California": "CA", "Colorado": "CO", "Connecticut": "CT",
        "Delaware": "DE", "Florida": "FL", "Georgia": "GA",
        "Hawaii": "HI", "Idaho": "ID", "Illinois": "IL", "Indiana": "IN",
        "Iowa": "IA", "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA",
        "Maine": "ME", "Maryland": "MD", "Massachusetts": "MA",
        "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS",
        "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV",
        "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM",
        "New York": "NY", "North Carolina": "NC", "North Dakota": "ND",
        "Ohio": "OH", "Oklahoma": "OK", "Oregon": "OR",
        "Pennsylvania": "PA", "Rhode Island": "RI",
        "South Carolina": "SC", "South Dakota": "SD",
        "Tennessee": "TN", "Texas": "TX", "Utah": "UT", "Vermont": "VT",
        "Virginia": "VA", "Washington": "WA",
        "West Virginia": "WV", "Wisconsin": "WI",
        "Wyoming": "WY", "District of Columbia": "DC",
    }

    def __init__(self, queryset):
        self.queryset = queryset

    # ------------------------------------------------------------------
    # 🔁 CLONED dehydration logic (safe, isolated)
    # ------------------------------------------------------------------
    def dehydrate_foreclosure(self, obj):
        row = {field: "-" for field in self.FIELDS}

        # ---- Core foreclosure fields
        row.update({
            "ID": obj.id,
            "Changed At": obj.changed_at,
            "Surplus Status": obj.surplus_status,
            "Sale Status": obj.sale_status,
            "State": obj.state,
            "County": obj.county,
            "Case Number": obj.case_number,
            "Sale Date": obj.sale_date,
            "Confirmation Date": getattr(obj, "confirmation_date", None),
            "Sale Type": obj.sale_type,
            "Judgment Amount": obj.fcl_final_judgment,
            "Sale Price": obj.sale_price,
            "Possible Surplus": obj.possible_surplus,
            "Verified Surplus": obj.verified_surplus,
            "Mailing Address": getattr(obj, "mailing_address", "-"),
        })

        # ---- Plaintiff
        plaintiffs = obj.plaintiff.all() if hasattr(obj, "plaintiff") else []
        row["Plaintiff"] = ", ".join(str(p) for p in plaintiffs) if plaintiffs else "-"

        # ---- Defendants
        defendants = obj.defendant.all() if hasattr(obj, "defendant") else []
        row["Defendant"] = ", ".join(str(d) for d in defendants) if defendants else "-"

        # ---- Property
        prop = obj.property.first() if hasattr(obj, "property") else None
        if prop:
            state_short = self.STATE_ABBREVIATIONS.get(
                (prop.state or "").strip().title(), "-"
            )
            row.update({
                "Parcel": prop.parcel or "-",
                "Street Address": prop.street_address or "-",
                "City": prop.city or "-",
                "ST": state_short,
                "Zip": prop.zip_code or "-",
            })

        # ---- Contacts (defendants + related)
        contacts = []
        for d in defendants:
            if d not in contacts:
                contacts.append(d)
            for rc in d.related_contacts.all():
                if rc not in contacts:
                    contacts.append(rc)

        contacts = contacts[:5]

        for i, contact in enumerate(contacts, start=1):
            row[f"Contact Name {i}"] = str(contact)
            emails = list(contact.emails.values_list("email_address", flat=True)[:1])
            row[f"Email {i}"] = emails[0] if emails else "-"

            wireless = list(contact.wireless.values_list("w_number", flat=True)[:4])
            landline = list(contact.landline.values_list("l_number", flat=True)[:4])

            for j in range(4):
                row[f"Wireless {i}.{j+1}"] = wireless[j] if j < len(wireless) else "-"
                row[f"Landline {i}.{j+1}"] = landline[j] if j < len(landline) else "-"

        return row

    # ------------------------------------------------------------------
    # DataFrame & Excel
    # ------------------------------------------------------------------
    def to_dataframe(self):
        data = [self.dehydrate_foreclosure(obj) for obj in self.queryset]
        df = pd.DataFrame(data, columns=self.FIELDS)

        # Remove timezone for Excel
        for col in df.select_dtypes(include=["datetimetz"]).columns:
            df[col] = df[col].dt.tz_localize(None)

        return df

    def export_to_excel(self, filename_prefix="dashboard_export"):
        df = self.to_dataframe()
        buffer = BytesIO()
        filename = f"{filename_prefix}_{timezone.now().date()}.xlsx"

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            sheet = "Dashboard Leads"
            df.to_excel(writer, index=False, sheet_name=sheet)

            ws = writer.sheets[sheet]

            header_fill = PatternFill("solid", start_color="516699")
            header_font = Font(color="FFFFFF", bold=True)
            header_align = Alignment(horizontal="center")

            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align

                max_len = max(
                    len(col_name),
                    *[len(str(v)) for v in df[col_name].astype(str).values[:200]]
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)

        buffer.seek(0)
        return filename, buffer, df
